import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Подключение к базе данных
conn = sqlite3.connect('users.db', check_same_thread=False)
cursor = conn.cursor()

def create_tables():
    cursor.execute('''CREATE TABLE IF NOT EXISTS users (
                        id INTEGER PRIMARY KEY,
                        user_id INTEGER UNIQUE,
                        name TEXT,
                        phone TEXT,
                        city TEXT
                    )''')
    cursor.execute('''CREATE TABLE IF NOT EXISTS requests (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER,
                        amount INTEGER,
                        comment TEXT,
                        file_path TEXT,
                        status TEXT DEFAULT 'В обробці',
                        request_number INTEGER,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,  -- Добавляем столбец для даты создания
                        FOREIGN KEY (user_id) REFERENCES users(user_id)
                    )''')
    conn.commit()

# Функция для проверки, зарегистрирован ли пользователь
def is_user_registered(user_id):
    cursor.execute("SELECT * FROM users WHERE user_id = ?", (user_id,))
    return cursor.fetchone() is not None

# Функция для добавления нового пользователя
def add_user(user_id, name, phone, city):
    cursor.execute("INSERT INTO users (user_id, name, phone, city) VALUES (?, ?, ?, ?)", (user_id, name, phone, city))
    conn.commit()

# Функция для добавления запроса
def add_request(user_id, amount, comment=None, file_path=None):
    request_number = get_next_request_number(user_id)
    cursor.execute("INSERT INTO requests (user_id, amount, comment, file_path, request_number) VALUES (?, ?, ?, ?, ?)", 
                   (user_id, amount, comment, file_path, request_number))
    conn.commit()
    request_id = cursor.lastrowid
    return request_id, request_number

# Функция для обновления статуса запроса
def update_request_status(request_id, status, comment=None):
    if comment:
        cursor.execute("UPDATE requests SET status = ?, comment = ? WHERE id = ?", (status, comment, request_id))
    else:
        cursor.execute("UPDATE requests SET status = ? WHERE id = ?", (status, request_id))
    conn.commit()

# Функция для получения информации о запросе
def get_request_info(request_id):
    cursor.execute("SELECT requests.*, users.name FROM requests JOIN users ON requests.user_id = users.user_id WHERE requests.id = ?", (request_id,))
    result = cursor.fetchone()
    if result:
        return {
            'id': result[0],
            'user_id': result[1],
            'amount': result[2],
            'comment': result[3],
            'file_path': result[4],
            'status': result[5],
            'request_number': result[6],
            'created_at': result[7],
            'name': result[8]
        }
    return None

# Функция для получения активных запросов пользователя
def get_active_requests(user_id):
    cursor.execute("SELECT request_number, amount, status FROM requests WHERE user_id = ? AND status = 'В обробці'", (user_id,))
    rows = cursor.fetchall()
    requests = []
    for row in rows:
        requests.append({
            'request_number': row[0],
            'amount': row[1],
            'status': row[2]
        })
    return requests

# Функция для получения информации о пользователе
def get_user_info(user_id):
    cursor.execute("SELECT name, phone, city FROM users WHERE user_id = ?", (user_id,))
    result = cursor.fetchone()
    if result:
        return {'name': result[0], 'phone': result[1], 'city': result[2]}
    return None

# Функция для получения следующего порядкового номера запроса для пользователя
def get_next_request_number(user_id):
    cursor.execute("SELECT COUNT(*) FROM requests WHERE user_id = ?", (user_id,))
    request_count = cursor.fetchone()[0]
    return request_count + 1

# Функция для получения всех активных запросов в системе
def get_all_active_requests():
    cursor.execute("SELECT request_number, name, phone, amount, comment, file_path, status, created_at FROM requests JOIN users ON requests.user_id = users.user_id WHERE status = 'В обробці'")
    rows = cursor.fetchall()
    active_requests = []
    for row in rows:
        active_requests.append({
            'request_number': row[0],
            'name': row[1],
            'phone': row[2],
            'amount': row[3],
            'comment': row[4],
            'file_path': row[5],
            'status': row[6],
            'created_at': row[7]
        })
    return active_requests

# Функция для получения общего количества запросов
def get_total_requests():
    cursor.execute("SELECT COUNT(*) FROM requests")
    return cursor.fetchone()[0]

# Функция для получения количества утвержденных запросов
def get_approved_requests():
    cursor.execute("SELECT COUNT(*) FROM requests WHERE status = 'Погоджено'")
    return cursor.fetchone()[0]

# Функция для получения общей суммы запросов
def get_total_amount():
    cursor.execute("SELECT SUM(amount) FROM requests")
    result = cursor.fetchone()
    return result[0] if result[0] else 0

# Функция для экспорта запросов в Excel с именами пользователей и датой создания
def export_requests_to_excel():
    cursor.execute('''SELECT 
            requests.id,
            users.name,
            users.phone,
            requests.amount,
            requests.comment,
            requests.file_path,
            requests.status,
            requests.request_number,
            requests.created_at  -- Добавляем дату создания
        FROM 
            requests 
        JOIN 
            users ON requests.user_id = users.user_id
    ''')
    rows = cursor.fetchall()
    
    # Определение названий столбцов
    columns = ['id', 'Имя', 'Телефон', 'Сумма', 'Комментарий', 'Файл', 'Статус', 'Номер запроса', 'Дата создания']  
    
    # Создание DataFrame
    df = pd.DataFrame(rows, columns=columns)
    
    # Путь к файлу Excel
    file_path = 'requests.xlsx'
    
    # Запись в файл Excel
    df.to_excel(file_path, index=False)

    # Открытие файла Excel для настройки ширины столбцов
    workbook = Workbook()
    worksheet = workbook.active

    for r in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(r)

    # Настройка ширины столбцов
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)  
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2  
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Сохранение файла
    workbook.save(file_path)

    return file_path

# Функция для получения количества отклоненных запросов
def get_rejected_requests():
    cursor.execute("SELECT COUNT(*) FROM requests WHERE status = 'Відхилено'")
    return cursor.fetchone()[0]

# Создание таблиц при импорте модуля
create_tables()
